"""
Live Mask Detection  —  v4 (Smart Per-Person Tracking)
=======================================================
Logic:
  - Each detected face gets a persistent ID based on position
  - A face is only logged ONCE per unique status (With Mask / No Mask)
  - If the SAME person changes status (puts on / removes mask), log again
  - If face disappears then reappears, treated as same person if in same region

Requirements:
    pip install ultralytics opencv-python openpyxl pillow requests numpy scipy
"""

import os, time, threading, shutil, tempfile, numpy as np
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import cv2
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    import requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

try:
    from ultralytics import YOLO
    YOLO_AVAILABLE = True
except ImportError:
    YOLO_AVAILABLE = False

EXCEL_FILE = "mask_detection_log.xlsx"

STATUS_COLOR = {
    "With Mask":     ("#27ae60", (0, 210, 0),   "🟢"),
    "No Mask":       ("#e74c3c", (0, 0, 220),   "🔴"),
    "Improper Mask": ("#f39c12", (0, 165, 255), "🟡"),
}

# ═══════════════════════════════════════════════════════════════════
# FACE TRACKER  — assigns stable IDs to faces across frames
# ═══════════════════════════════════════════════════════════════════

class FaceTracker:
    """
    Lightweight centroid tracker.
    - Matches detections to existing tracks by distance.
    - Each track remembers its last logged status.
    - Only triggers a log when status CHANGES for that track.
    """

    MAX_DISAPPEARED = 40    # frames before a track is removed
    MAX_DISTANCE    = 120   # pixels — max centroid shift to match same face

    def __init__(self):
        self.next_id    = 0
        self.tracks     = {}   # id → {centroid, status, disappeared, bbox}
        self._lock      = threading.Lock()

    def _centroid(self, x1, y1, x2, y2):
        return np.array([(x1+x2)//2, (y1+y2)//2], dtype=float)

    def update(self, detections):
        """
        detections: list of (x1,y1,x2,y2,status,conf)
        Returns:    list of (x1,y1,x2,y2,status,conf,face_id,is_new_status)
        """
        with self._lock:
            # ── mark all tracks as disappeared this frame ──────────
            for tid in list(self.tracks):
                self.tracks[tid]["disappeared"] += 1

            if not detections:
                # purge stale
                for tid in [t for t,v in self.tracks.items()
                            if v["disappeared"] > self.MAX_DISAPPEARED]:
                    del self.tracks[tid]
                return []

            centroids = [self._centroid(*d[:4]) for d in detections]
            output    = []

            if not self.tracks:
                # Register all as new
                for i, det in enumerate(detections):
                    tid = self._register(centroids[i], det)
                    x1,y1,x2,y2,status,conf = det
                    output.append((x1,y1,x2,y2,status,conf,tid, True))
                return output

            # Match detections → existing tracks by nearest centroid
            track_ids   = list(self.tracks.keys())
            track_cents = np.array([self.tracks[t]["centroid"]
                                    for t in track_ids])

            used_tracks = set()
            used_dets   = set()

            # distance matrix
            diffs = (track_cents[:, None, :] -
                     np.array(centroids)[None, :, :])
            dists = np.linalg.norm(diffs, axis=2)   # shape (T, D)

            # greedy nearest-neighbour match
            for _ in range(min(len(track_ids), len(detections))):
                if dists.size == 0:
                    break
                t_idx, d_idx = np.unravel_index(dists.argmin(), dists.shape)
                if dists[t_idx, d_idx] > self.MAX_DISTANCE:
                    break
                tid = track_ids[t_idx]
                used_tracks.add(t_idx)
                used_dets.add(d_idx)
                det    = detections[d_idx]
                status = det[4]
                # Update track
                self.tracks[tid]["centroid"]    = centroids[d_idx]
                self.tracks[tid]["bbox"]        = det[:4]
                self.tracks[tid]["disappeared"] = 0
                # Determine if status changed
                prev   = self.tracks[tid]["last_logged_status"]
                is_new = (prev != status)
                if is_new:
                    self.tracks[tid]["last_logged_status"] = status
                output.append((*det, tid, is_new))
                # blank out row/col so it can't be reused
                dists[t_idx, :] = 1e9
                dists[:, d_idx] = 1e9

            # Unmatched detections → new tracks
            for d_idx, det in enumerate(detections):
                if d_idx not in used_dets:
                    tid = self._register(centroids[d_idx], det)
                    output.append((*det, tid, True))

            # Purge stale tracks
            for tid in [t for t,v in self.tracks.items()
                        if v["disappeared"] > self.MAX_DISAPPEARED]:
                del self.tracks[tid]

            return output

    def _register(self, centroid, det):
        tid = self.next_id
        self.next_id += 1
        self.tracks[tid] = {
            "centroid":            centroid,
            "bbox":                det[:4],
            "disappeared":         0,
            "last_logged_status":  det[4],   # log immediately on first sight
        }
        return tid


# ═══════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════

def _excel_init(path):
    if os.path.exists(path):
        try:
            return openpyxl.load_workbook(path)
        except Exception:
            pass
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detection Log"
    headers = ["#", "Timestamp", "Date", "Time",
               "Face ID", "Status", "Confidence (%)", "Notes"]
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for i, w in enumerate([5, 22, 12, 10, 8, 18, 16, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return wb


def excel_log(face_id: int, status: str, conf: float, notes: str = ""):
    wb  = _excel_init(EXCEL_FILE)
    ws  = wb.active
    row = ws.max_row + 1
    now = datetime.now()

    fills = {"With Mask":     ("C6EFCE", "276221"),
             "No Mask":       ("FFC7CE", "9C0006"),
             "Improper Mask": ("FFEB9C", "9C5700")}
    bg, fg = fills.get(status, ("FFFFFF", "000000"))
    fill   = PatternFill("solid", fgColor=bg)
    font   = Font(color=fg, name="Arial")

    vals = [row-1,
            now.strftime("%Y-%m-%d %H:%M:%S"),
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            f"Face-{face_id}",
            status,
            round(conf*100, 1),
            notes]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill, cell.font = fill, font
        cell.alignment = Alignment(horizontal="center")

    # Safe save via temp file
    tmp_fd, tmp_path = tempfile.mkstemp(
        suffix=".xlsx",
        dir=os.path.dirname(os.path.abspath(EXCEL_FILE)) or ".")
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
        shutil.move(tmp_path, EXCEL_FILE)
    except PermissionError:
        try: os.remove(tmp_path)
        except: pass
        raise


# ═══════════════════════════════════════════════════════════════════
# SKIN HEURISTIC
# ═══════════════════════════════════════════════════════════════════

_SKIN_LO = np.array([0,  20,  70], np.uint8)
_SKIN_HI = np.array([20, 255, 255], np.uint8)

def _skin_ratio(roi):
    if roi is None or roi.size == 0:
        return 0.0
    hsv  = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    mask = cv2.inRange(hsv, _SKIN_LO, _SKIN_HI)
    return cv2.countNonZero(mask) / max(1, roi.shape[0]*roi.shape[1])

def classify_face(frame, x1, y1, x2, y2):
    fh    = y2-y1
    y_mid = y1+fh//2
    if _skin_ratio(frame[y_mid:y2, x1:x2]) > 0.55:
        return "No Mask", 0.82
    y_nose = y1+int(fh*0.45)
    y_lip  = y1+int(fh*0.65)
    if _skin_ratio(frame[y_nose:y_lip, x1:x2]) > 0.40:
        return "Improper Mask", 0.74
    return "With Mask", 0.86


# ═══════════════════════════════════════════════════════════════════
# DETECTION BACKENDS
# ═══════════════════════════════════════════════════════════════════

class HaarDetector:
    def __init__(self):
        self.cc = cv2.CascadeClassifier(
            cv2.data.haarcascades+"haarcascade_frontalface_default.xml")
    def detect(self, frame):
        gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = self.cc.detectMultiScale(
                    gray, 1.1, 5, minSize=(60,60))
        return [(fx,fy,fx+fw,fy+fh)+classify_face(frame,fx,fy,fx+fw,fy+fh)
                for (fx,fy,fw,fh) in faces]

FACE_MODEL_URL  = ("https://github.com/ultralytics/assets/releases/"
                   "download/v0.0.0/yolov8n-face.pt")
FACE_MODEL_PATH = "yolov8n-face.pt"

class YOLOFaceDetector:
    def __init__(self, model_path, conf=0.45):
        self.model = YOLO(model_path)
        self.conf  = conf
    def detect(self, frame):
        out = []
        for r in self.model(frame, verbose=False, conf=self.conf):
            for box in r.boxes:
                x1,y1,x2,y2 = map(int, box.xyxy[0])
                out.append((x1,y1,x2,y2)+classify_face(frame,x1,y1,x2,y2))
        return out

_LABEL_MAP = {
    "with_mask":"With Mask","mask_weared_correct":"With Mask",
    "mask":"With Mask","wearing_mask":"With Mask",
    "without_mask":"No Mask","no_mask":"No Mask",
    "mask_weared_incorrect":"Improper Mask",
    "improper_mask":"Improper Mask","incorrect_mask":"Improper Mask",
}

class YOLOMaskDetector:
    def __init__(self, model_path, conf=0.45):
        self.model = YOLO(model_path)
        self.conf  = conf
    def detect(self, frame):
        out = []
        for r in self.model(frame, verbose=False, conf=self.conf):
            for box in r.boxes:
                x1,y1,x2,y2 = map(int, box.xyxy[0])
                conf   = float(box.conf[0])
                label  = self.model.names.get(
                    int(box.cls[0]),"").lower().replace(" ","_")
                out.append((x1,y1,x2,y2,
                            _LABEL_MAP.get(label,f"Class:{label}"),conf))
        return out


# ═══════════════════════════════════════════════════════════════════
# APPLICATION
# ═══════════════════════════════════════════════════════════════════

class App:
    def __init__(self, root):
        self.root        = root
        self.root.title("🎭 Live Mask Detection")
        self.root.configure(bg="#1a1a2e")
        self.root.resizable(True, True)
        self.cap         = None
        self.detector    = None
        self.tracker     = FaceTracker()
        self.running     = False
        self.log_int     = 3
        self.stats       = {"With Mask":0,"No Mask":0,
                            "Improper Mask":0,"total":0}
        self._imgtk      = None
        self._excel_warn = False
        self._build_ui()
        threading.Thread(target=self._init_detector, daemon=True).start()

    # ── UI ──────────────────────────────────────────────────────────

    def _build_ui(self):
        hdr = tk.Frame(self.root, bg="#16213e", pady=8)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🎭  Live Mask Detection System",
                 font=("Arial",18,"bold"), fg="#e94560", bg="#16213e").pack()
        tk.Label(hdr,
                 text="Each person logged ONCE per status change · Excel auto-save",
                 font=("Arial",9), fg="#aaaaaa", bg="#16213e").pack()

        body = tk.Frame(self.root, bg="#1a1a2e")
        body.pack(fill="both", expand=True, padx=8, pady=8)

        # LEFT
        lf = tk.Frame(body, bg="#1a1a2e")
        lf.pack(side="left", fill="both", expand=True)
        self.canvas = tk.Canvas(lf, bg="#0d1b2a",
                                highlightthickness=2,
                                highlightbackground="#2c3e50")
        self.canvas.pack(fill="both", expand=True, padx=4, pady=4)
        self.badge_var = tk.StringVar(value="⬤  Idle")
        self.badge_lbl = tk.Label(lf, textvariable=self.badge_var,
                                  font=("Arial",16,"bold"),
                                  fg="#aaaaaa", bg="#1a1a2e")
        self.badge_lbl.pack(pady=6)

        # RIGHT
        rp = tk.Frame(body, bg="#16213e", width=280, padx=14, pady=14)
        rp.pack(side="right", fill="y", padx=(6,0))
        rp.pack_propagate(False)

        def sep(lbl):
            tk.Label(rp, text=lbl, font=("Arial",11,"bold"),
                     fg="#e94560", bg="#16213e").pack(anchor="w", pady=(10,0))
            ttk.Separator(rp, orient="horizontal").pack(fill="x", pady=3)

        def kv(lbl, wfn):
            f = tk.Frame(rp, bg="#16213e"); f.pack(fill="x", pady=2)
            tk.Label(f, text=lbl, fg="white", bg="#16213e",
                     font=("Arial",9), width=19, anchor="w").pack(side="left")
            wfn(f)

        sep("⚙  Controls")
        self.cam_var  = tk.IntVar(value=0)
        self.lint_var = tk.IntVar(value=3)
        self.conf_var = tk.DoubleVar(value=0.45)

        kv("Camera index:", lambda f: tk.Spinbox(
            f, from_=0, to=5, textvariable=self.cam_var,
            width=5, font=("Arial",9)).pack(side="right"))
        kv("Log interval (s):", lambda f: tk.Spinbox(
            f, from_=1, to=60, textvariable=self.lint_var,
            width=5, font=("Arial",9)).pack(side="right"))
        kv("Confidence:", lambda f: tk.Scale(
            f, variable=self.conf_var, from_=0.1, to=0.95,
            resolution=0.05, orient="horizontal", bg="#16213e",
            fg="white", length=100,
            highlightthickness=0).pack(side="right"))

        bs = dict(font=("Arial",10,"bold"), width=22,
                  relief="flat", cursor="hand2", pady=7)
        self.btn_start = tk.Button(rp, text="▶  Start Detection",
                                   bg="#27ae60", fg="white",
                                   command=self.start, **bs)
        self.btn_start.pack(pady=(12,3))
        self.btn_stop = tk.Button(rp, text="⏹  Stop",
                                  bg="#e74c3c", fg="white",
                                  command=self.stop,
                                  state="disabled", **bs)
        self.btn_stop.pack(pady=3)
        tk.Button(rp, text="📊  Open Excel Log",
                  bg="#2980b9", fg="white",
                  command=self.open_excel, **bs).pack(pady=3)
        tk.Button(rp, text="🔄  Reset Tracker",
                  bg="#7f8c8d", fg="white",
                  command=self.reset_tracker, **bs).pack(pady=3)

        self.excel_sv = tk.StringVar(value="📁 Excel: Ready")
        tk.Label(rp, textvariable=self.excel_sv, fg="#2ecc71",
                 bg="#16213e", font=("Arial",8)).pack(anchor="w", pady=(2,0))

        sep("📈  Session Stats")
        self._svars = {}
        for k, (_,_,emoji) in STATUS_COLOR.items():
            sv = tk.StringVar(value=f"{emoji}  {k}: 0")
            self._svars[k] = sv
            tk.Label(rp, textvariable=sv, fg="white",
                     bg="#16213e", font=("Arial",10)).pack(anchor="w", pady=1)
        self._total_sv = tk.StringVar(value="📋  Total logged: 0")
        tk.Label(rp, textvariable=self._total_sv, fg="white",
                 bg="#16213e", font=("Arial",10)).pack(anchor="w", pady=1)

        # Active face count
        self._faces_sv = tk.StringVar(value="👤  Active faces: 0")
        tk.Label(rp, textvariable=self._faces_sv, fg="#3498db",
                 bg="#16213e", font=("Arial",10,"bold")).pack(anchor="w",
                                                               pady=1)

        sep("🤖  Model Status")
        self.model_sv = tk.StringVar(value="⏳ Initialising…")
        tk.Label(rp, textvariable=self.model_sv, fg="#f39c12",
                 bg="#16213e", font=("Arial",9),
                 wraplength=235, justify="left").pack(anchor="w")

        sep("📋  Recent Logs")
        self.log_box = tk.Text(rp, height=7, bg="#0d1b2a", fg="white",
                               font=("Courier",8), state="disabled",
                               relief="flat", wrap="none")
        self.log_box.pack(fill="x")

    # ── Detector init ───────────────────────────────────────────────

    def _init_detector(self):
        if YOLO_AVAILABLE and os.path.exists("mask_yolov8.pt"):
            try:
                self.detector = YOLOMaskDetector("mask_yolov8.pt",
                                                  self.conf_var.get())
                self.model_sv.set("✅ Custom mask_yolov8.pt")
                return
            except Exception:
                pass

        if YOLO_AVAILABLE and REQUESTS_OK:
            try:
                if not os.path.exists(FACE_MODEL_PATH):
                    self.model_sv.set("⬇ Downloading YOLOv8-face (~6 MB)…")
                    r = requests.get(FACE_MODEL_URL, timeout=40, stream=True)
                    if r.status_code == 200:
                        with open(FACE_MODEL_PATH,"wb") as f:
                            for chunk in r.iter_content(65536):
                                f.write(chunk)
                self.detector = YOLOFaceDetector(FACE_MODEL_PATH,
                                                   self.conf_var.get())
                self.model_sv.set(
                    "✅ YOLOv8-face + heuristic\n"
                    "Place mask_yolov8.pt here for best accuracy")
                return
            except Exception:
                pass

        self.detector = HaarDetector()
        self.model_sv.set(
            "✅ OpenCV Haar + mask heuristic\n"
            "Place mask_yolov8.pt here for best accuracy")

    # ── Camera ──────────────────────────────────────────────────────

    def start(self):
        cam = self.cam_var.get()
        cap = (cv2.VideoCapture(cam, cv2.CAP_DSHOW)
               if os.name == "nt" else cv2.VideoCapture(cam))
        if not cap.isOpened():
            cap = cv2.VideoCapture(cam)
        if not cap.isOpened():
            messagebox.showerror("Camera Error",
                                 f"Cannot open camera {cam}.")
            return
        cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
        cap.set(cv2.CAP_PROP_FPS, 30)
        self.cap          = cap
        self.running      = True
        self.log_int      = self.lint_var.get()
        self._excel_warn  = False
        self.tracker      = FaceTracker()   # fresh tracker on each start
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        threading.Thread(target=self._loop, daemon=True).start()

    def stop(self):
        self.running = False
        if self.cap:
            self.cap.release()
            self.cap = None
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.canvas.delete("all")
        self.badge_var.set("⬤  Stopped")
        self.badge_lbl.config(fg="#aaaaaa")

    def reset_tracker(self):
        """Forget all face IDs — next detection treated as new."""
        self.tracker = FaceTracker()
        self.root.after(0, self.excel_sv.set,
                        "🔄 Tracker reset — faces will re-log")

    # ── Detection loop ──────────────────────────────────────────────

    def _loop(self):
        while self.running:
            try:
                ret, frame = self.cap.read()
                if not ret:
                    time.sleep(0.03); continue

                annotated = frame.copy()

                # Raw detections
                try:
                    raw = self.detector.detect(frame) if self.detector else []
                except Exception:
                    raw = []

                # Feed through tracker
                tracked = self.tracker.update(raw)

                best_s, best_c = "Unknown", 0.0

                for (x1, y1, x2, y2, status, conf,
                     face_id, is_new_status) in tracked:

                    _, bgr, emoji = STATUS_COLOR.get(
                        status, ("#aaa",(150,150,150),""))

                    # Draw bounding box
                    cv2.rectangle(annotated, (x1,y1), (x2,y2), bgr, 3)

                    # Label: Face-ID + status + conf
                    lbl = f"  Face-{face_id} | {status}  {conf:.0%}"
                    (tw,th),_ = cv2.getTextSize(
                        lbl, cv2.FONT_HERSHEY_SIMPLEX, 0.65, 2)
                    cv2.rectangle(annotated,
                                  (x1, y1-th-16),(x1+tw+4, y1), bgr, -1)
                    cv2.putText(annotated, lbl, (x1+2, y1-6),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.65,
                                (255,255,255), 2, cv2.LINE_AA)

                    # "NEW" badge if just logged
                    if is_new_status:
                        cv2.putText(annotated, "● LOGGED",
                                    (x1, y2+22),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.55,
                                    bgr, 2, cv2.LINE_AA)
                        self._record(face_id, status, conf)

                    if conf > best_c:
                        best_c, best_s = conf, status

                # Update face count indicator
                active = len(self.tracker.tracks)
                self.root.after(0, self._faces_sv.set,
                                f"👤  Active faces: {active}")

                self.root.after(0, self._render_frame, annotated.copy())
                self.root.after(0, self._badge, best_s, best_c)

            except Exception as e:
                print(f"[loop error] {e}")
                time.sleep(0.1)

    # ── Render ──────────────────────────────────────────────────────

    def _render_frame(self, frame):
        if not self.running: return
        cw = max(self.canvas.winfo_width(),  640)
        ch = max(self.canvas.winfo_height(), 440)
        rgb   = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h,w   = rgb.shape[:2]
        sc    = min(cw/w, ch/h)
        nw,nh = int(w*sc), int(h*sc)
        img   = Image.fromarray(rgb).resize((nw,nh), Image.LANCZOS)
        imgtk = ImageTk.PhotoImage(image=img)
        self.canvas.delete("all")
        self.canvas.create_image(cw//2, ch//2, anchor="center", image=imgtk)
        self._imgtk = imgtk

    def _badge(self, s, c):
        if s in STATUS_COLOR:
            col,_,emoji = STATUS_COLOR[s]
            self.badge_var.set(f"{emoji}  {s.upper()}  ({c:.0%})")
            self.badge_lbl.config(fg=col)
        else:
            self.badge_var.set("⬤  No face detected")
            self.badge_lbl.config(fg="#aaaaaa")

    # ── Record ──────────────────────────────────────────────────────

    def _record(self, face_id, status, conf):
        self.stats["total"] += 1
        if status in self.stats:
            self.stats[status] += 1

        try:
            excel_log(face_id, status, conf,
                      f"Status change detected")
            self.root.after(0, self.excel_sv.set, "📁 Excel: Saved ✓")
        except PermissionError:
            self.root.after(0, self.excel_sv.set,
                            "⚠ Excel locked — close the file!")
            if not self._excel_warn:
                self._excel_warn = True
                self.root.after(0, messagebox.showwarning,
                    "Excel File Locked",
                    "Close mask_detection_log.xlsx in Excel\n"
                    "so the app can write to it.")
        except Exception as e:
            self.root.after(0, self.excel_sv.set, f"⚠ {e}")

        def _upd():
            for k,sv in self._svars.items():
                _,_,emoji = STATUS_COLOR[k]
                sv.set(f"{emoji}  {k}: {self.stats[k]}")
            self._total_sv.set(
                f"📋  Total logged: {self.stats['total']}")
            entry = (f"[{datetime.now().strftime('%H:%M:%S')}] "
                     f"Face-{face_id:<3} {status:<18} {conf:.0%}\n")
            self.log_box.config(state="normal")
            self.log_box.insert("end", entry)
            self.log_box.see("end")
            self.log_box.config(state="disabled")

        self.root.after(0, _upd)

    # ── Misc ────────────────────────────────────────────────────────

    def open_excel(self):
        if not os.path.exists(EXCEL_FILE):
            wb = _excel_init(EXCEL_FILE)
            wb.save(EXCEL_FILE)
        (os.startfile(EXCEL_FILE) if os.name=="nt"
         else os.system(f"xdg-open '{EXCEL_FILE}'"))

    def on_close(self):
        self.running = False
        if self.cap: self.cap.release()
        self.root.destroy()


# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    app  = App(root)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.geometry("1150x680")
    root.minsize(920, 560)
    root.mainloop()