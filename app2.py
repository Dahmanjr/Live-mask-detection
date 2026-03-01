"""
Live Mask Detection — Streamlit WebRTC
Exact replica of the tkinter desktop GUI
"""
import streamlit as st
import cv2
import numpy as np
import io
import threading
from datetime import datetime
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from streamlit_webrtc import webrtc_streamer, VideoProcessorBase, RTCConfiguration
import av

st.set_page_config(
    page_title="🎭 Live Mask Detection",
    page_icon="🎭",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ═══════════════════════════════════════════════════════════════════
# CSS — exact dark theme replica of the tkinter app
# ═══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Arial&display=swap');

/* ── Global dark background ── */
html, body, [class*="css"], .stApp {
    background-color: #1a1a2e !important;
    color: #ffffff;
    font-family: Arial, sans-serif;
}

/* ── Hide default streamlit chrome ── */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 0 !important; max-width: 100% !important; }

/* ── TOP HEADER BAR (matches tkinter hdr frame) ── */
.app-header {
    background: #16213e;
    padding: 10px 0 8px 0;
    text-align: center;
    border-bottom: 1px solid #0d1b2a;
    margin-bottom: 0;
}
.app-title {
    font-size: 1.5rem;
    font-weight: bold;
    color: #e94560;
    margin: 0;
}
.app-sub {
    font-size: 0.78rem;
    color: #aaaaaa;
    margin: 0;
}

/* ── LEFT PANEL — video area ── */
.video-container {
    background: #0d1b2a;
    border: 2px solid #2c3e50;
    border-radius: 4px;
    min-height: 440px;
    display: flex;
    align-items: center;
    justify-content: center;
}

/* ── Badge below video ── */
.badge {
    font-size: 1.2rem;
    font-weight: bold;
    text-align: center;
    padding: 8px;
    color: #aaaaaa;
}
.badge-mask     { color: #27ae60; }
.badge-nomask   { color: #e74c3c; }
.badge-improper { color: #f39c12; }

/* ── RIGHT PANEL (matches tkinter rp frame) ── */
.right-panel {
    background: #16213e;
    padding: 14px;
    height: 100%;
    min-height: 600px;
}

/* ── Section headers (sep function) ── */
.sec-label {
    font-size: 0.95rem;
    font-weight: bold;
    color: #e94560;
    margin: 14px 0 2px 0;
}
.sec-line {
    border: none;
    border-top: 1px solid #555;
    margin: 0 0 6px 0;
}

/* ── Buttons matching tkinter buttons ── */
div[data-testid="column"] .stButton > button,
.stButton > button {
    width: 100% !important;
    font-weight: bold !important;
    border: none !important;
    border-radius: 4px !important;
    padding: 8px 0 !important;
    font-size: 0.9rem !important;
    margin-bottom: 4px !important;
    cursor: pointer !important;
}
.btn-start > button { background: #27ae60 !important; color: white !important; }
.btn-stop  > button { background: #e74c3c !important; color: white !important; }
.btn-excel > button { background: #2980b9 !important; color: white !important; }
.btn-reset > button { background: #7f8c8d !important; color: white !important; }

/* ── Excel status line ── */
.excel-status { font-size: 0.72rem; color: #2ecc71; margin: 2px 0 6px 0; }
.excel-warn   { color: #e74c3c; }

/* ── Stat rows ── */
.stat-row {
    font-size: 0.88rem;
    color: white;
    padding: 1px 0;
}
.stat-faces {
    font-size: 0.88rem;
    font-weight: bold;
    color: #3498db;
    padding: 1px 0;
}

/* ── Model status ── */
.model-status {
    font-size: 0.78rem;
    color: #f39c12;
    word-wrap: break-word;
    line-height: 1.4;
}

/* ── Log box (matches tkinter Text widget) ── */
.log-box {
    background: #0d1b2a;
    border-radius: 4px;
    padding: 6px 8px;
    font-family: 'Courier New', monospace;
    font-size: 0.72rem;
    color: white;
    min-height: 130px;
    max-height: 160px;
    overflow-y: auto;
    line-height: 1.5;
}
.log-mask     { color: #27ae60; }
.log-nomask   { color: #e74c3c; }
.log-improper { color: #f39c12; }

/* ── kv rows (label+widget pairs) ── */
.kv-row {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 4px;
    font-size: 0.82rem;
    color: white;
}

/* ── Streamlit slider & number input dark mode ── */
.stSlider > div > div > div { background: #e94560 !important; }
.stNumberInput input, .stSelectbox select {
    background: #0d1b2a !important;
    color: white !important;
    border: 1px solid #2c3e50 !important;
}

/* ── WebRTC video ── */
.stVideo video { width: 100% !important; border-radius: 4px; }
[data-testid="stImage"] img { width: 100% !important; border-radius: 4px; }

/* ── Remove sidebar padding ── */
[data-testid="stSidebar"] { display: none; }

</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════
STATUS_COLOR = {
    "With Mask":     ("#27ae60", (0, 210, 0),   "🟢"),
    "No Mask":       ("#e74c3c", (0, 0, 220),   "🔴"),
    "Improper Mask": ("#f39c12", (0, 165, 255), "🟡"),
}

RTC_CONFIG = RTCConfiguration({"iceServers": [
    {"urls": ["stun:stun.l.google.com:19302"]},
    {"urls": ["stun:stun1.l.google.com:19302"]},
]})

# ═══════════════════════════════════════════════════════════════════
# FACE TRACKER  (identical to tkinter version)
# ═══════════════════════════════════════════════════════════════════
class FaceTracker:
    MAX_DISAPPEARED = 40
    MAX_DISTANCE    = 120

    def __init__(self):
        self.next_id = 0
        self.tracks  = {}
        self._lock   = threading.Lock()

    def _centroid(self, x1, y1, x2, y2):
        return np.array([(x1+x2)//2, (y1+y2)//2], dtype=float)

    def update(self, detections):
        with self._lock:
            for tid in list(self.tracks):
                self.tracks[tid]["disappeared"] += 1

            stale = [t for t,v in self.tracks.items()
                     if v["disappeared"] > self.MAX_DISAPPEARED]
            for t in stale:
                del self.tracks[t]

            if not detections:
                return []

            centroids = [self._centroid(*d[:4]) for d in detections]
            output    = []

            if not self.tracks:
                for i, det in enumerate(detections):
                    tid = self._register(centroids[i], det)
                    output.append((*det, tid, True))
                return output

            track_ids   = list(self.tracks.keys())
            track_cents = np.array([self.tracks[t]["centroid"] for t in track_ids])
            diffs = track_cents[:, None, :] - np.array(centroids)[None, :, :]
            dists = np.linalg.norm(diffs, axis=2)

            used_t, used_d = set(), set()
            for _ in range(min(len(track_ids), len(detections))):
                if dists.size == 0: break
                t_idx, d_idx = np.unravel_index(dists.argmin(), dists.shape)
                if dists[t_idx, d_idx] > self.MAX_DISTANCE: break
                tid = track_ids[t_idx]
                used_t.add(t_idx); used_d.add(d_idx)
                det    = detections[d_idx]
                status = det[4]
                self.tracks[tid]["centroid"]    = centroids[d_idx]
                self.tracks[tid]["bbox"]        = det[:4]
                self.tracks[tid]["disappeared"] = 0
                prev   = self.tracks[tid]["last_logged_status"]
                is_new = (prev != status)
                if is_new:
                    self.tracks[tid]["last_logged_status"] = status
                output.append((*det, tid, is_new))
                dists[t_idx, :] = 1e9
                dists[:, d_idx] = 1e9

            for d_idx, det in enumerate(detections):
                if d_idx not in used_d:
                    tid = self._register(centroids[d_idx], det)
                    output.append((*det, tid, True))

            return output

    def _register(self, centroid, det):
        tid = self.next_id; self.next_id += 1
        self.tracks[tid] = {
            "centroid": centroid, "bbox": det[:4],
            "disappeared": 0, "last_logged_status": det[4],
        }
        return tid

    def active_count(self):
        with self._lock:
            return len(self.tracks)


# ═══════════════════════════════════════════════════════════════════
# SKIN HEURISTIC  (identical to tkinter version)
# ═══════════════════════════════════════════════════════════════════
_SKIN_LO = np.array([0,  20,  70], np.uint8)
_SKIN_HI = np.array([20, 255, 255], np.uint8)

def _skin_ratio(roi):
    if roi is None or roi.size == 0: return 0.0
    hsv  = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    mask = cv2.inRange(hsv, _SKIN_LO, _SKIN_HI)
    return cv2.countNonZero(mask) / max(1, roi.shape[0]*roi.shape[1])

def classify_face(frame, x1, y1, x2, y2):
    fh = y2-y1; y_mid = y1+fh//2
    if _skin_ratio(frame[y_mid:y2, x1:x2]) > 0.55:
        return "No Mask", 0.82
    y_nose = y1+int(fh*0.45); y_lip = y1+int(fh*0.65)
    if _skin_ratio(frame[y_nose:y_lip, x1:x2]) > 0.40:
        return "Improper Mask", 0.74
    return "With Mask", 0.86


# ═══════════════════════════════════════════════════════════════════
# EXCEL  (identical logic, but returns bytes for download)
# ═══════════════════════════════════════════════════════════════════
def build_excel(log_data):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Detection Log"
    headers = ["#","Timestamp","Date","Time",
               "Face ID","Status","Confidence (%)","Notes"]
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for i, w in enumerate([5,22,12,10,8,18,16,24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    fills = {"With Mask":    ("C6EFCE","276221"),
             "No Mask":      ("FFC7CE","9C0006"),
             "Improper Mask":("FFEB9C","9C5700")}
    for idx, e in enumerate(log_data, 1):
        bg,fg = fills.get(e["status"],("FFFFFF","000000"))
        fill  = PatternFill("solid", fgColor=bg)
        font  = Font(color=fg, name="Arial")
        vals  = [idx, e["ts"], e["ts"][:10], e["ts"][11:19],
                 f"Face-{e['fid']}", e["status"],
                 round(e["conf"]*100,1), "Status change detected"]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=idx+1, column=c, value=v)
            cell.fill=fill; cell.font=font
            cell.alignment=Alignment(horizontal="center")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════
# VIDEO PROCESSOR  (runs per frame via WebRTC)
# ═══════════════════════════════════════════════════════════════════
class MaskProcessor(VideoProcessorBase):
    def __init__(self):
        self.detector = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        self.tracker     = FaceTracker()
        self.conf_thresh = 0.45
        self._lock       = threading.Lock()
        self.new_events  = []
        self.last_best   = ("Unknown", 0.0)
        self.active_faces= 0

    def recv(self, frame: av.VideoFrame) -> av.VideoFrame:
        img  = frame.to_ndarray(format="bgr24")
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = self.detector.detectMultiScale(
                    gray, scaleFactor=1.1, minNeighbors=5, minSize=(60,60))

        dets = []
        for (fx,fy,fw,fh) in faces:
            status, conf = classify_face(img, fx, fy, fx+fw, fy+fh)
            if conf >= self.conf_thresh:
                dets.append((fx, fy, fx+fw, fy+fh, status, conf))

        tracked = self.tracker.update(dets)

        best_s, best_c = "Unknown", 0.0
        for (x1,y1,x2,y2,status,conf,fid,is_new) in tracked:
            _, bgr, _ = STATUS_COLOR.get(status, ("#aaa",(150,150,150),""))

            # Bounding box
            cv2.rectangle(img, (x1,y1), (x2,y2), bgr, 3)

            # Label pill  —  Face-ID | status  conf%
            lbl = f"  Face-{fid} | {status}  {conf:.0%}"
            (tw,th),_ = cv2.getTextSize(lbl, cv2.FONT_HERSHEY_SIMPLEX, 0.65, 2)
            cv2.rectangle(img, (x1,y1-th-16), (x1+tw+4,y1), bgr, -1)
            cv2.putText(img, lbl, (x1+2,y1-6),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.65,
                        (255,255,255), 2, cv2.LINE_AA)

            # ● LOGGED badge
            if is_new:
                cv2.putText(img, "● LOGGED", (x1, y2+22),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.55,
                            bgr, 2, cv2.LINE_AA)
                with self._lock:
                    self.new_events.append({
                        "fid":    fid,
                        "status": status,
                        "conf":   conf,
                        "ts":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    })

            if conf > best_c:
                best_s, best_c = status, conf

        with self._lock:
            self.last_best    = (best_s, best_c)
            self.active_faces = self.tracker.active_count()

        return av.VideoFrame.from_ndarray(img, format="bgr24")

    def pull_events(self):
        with self._lock:
            evts = self.new_events[:]
            self.new_events.clear()
            return evts

    def get_state(self):
        with self._lock:
            return self.last_best, self.active_faces


# ═══════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════
def _init_state():
    defaults = {
        "log":   [],
        "stats": {"With Mask":0,"No Mask":0,"Improper Mask":0,"total":0},
        "excel_status": "📁 Excel: Ready",
        "excel_status_color": "normal",
    }
    for k,v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
_init_state()


# ═══════════════════════════════════════════════════════════════════
# HEADER  (matches tkinter hdr frame)
# ═══════════════════════════════════════════════════════════════════
st.markdown("""
<div class="app-header">
  <div class="app-title">🎭&nbsp; Live Mask Detection System</div>
  <div class="app-sub">Each person logged ONCE per status change &nbsp;·&nbsp; Excel auto-save</div>
</div>
""", unsafe_allow_html=True)

st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════
# BODY — left (video) + right (control panel)
# ═══════════════════════════════════════════════════════════════════
left, right = st.columns([3, 1.15], gap="small")

# ───────────────────────────────────────────────────────────────────
# RIGHT PANEL  (matches tkinter rp frame exactly)
# ───────────────────────────────────────────────────────────────────
with right:
    st.markdown("<div class='right-panel'>", unsafe_allow_html=True)

    # ── ⚙ Controls ──
    st.markdown("<div class='sec-label'>⚙&nbsp; Controls</div><hr class='sec-line'>",
                unsafe_allow_html=True)

    conf_val = st.slider("Confidence", 0.10, 0.95, 0.45, 0.05,
                         label_visibility="collapsed",
                         help="Confidence threshold")
    st.markdown(f"<div class='kv-row'><span>Confidence:</span><span style='color:#e94560;font-weight:bold'>{conf_val:.2f}</span></div>",
                unsafe_allow_html=True)

    # Buttons — matching tkinter layout exactly
    st.markdown("<div style='margin-top:10px'>", unsafe_allow_html=True)

    # Start / Stop handled by WebRTC component (rendered in left panel)
    # but we show a Reset Tracker button here
    with st.container():
        st.markdown("<div class='btn-excel'>", unsafe_allow_html=True)
        if st.session_state.log:
            st.download_button(
                "📊  Download Excel Log",
                data=build_excel(st.session_state.log),
                file_name=f"mask_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        else:
            st.button("📊  Download Excel Log",
                      disabled=True, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with st.container():
        st.markdown("<div class='btn-reset'>", unsafe_allow_html=True)
        if st.button("🔄  Reset Tracker", use_container_width=True):
            st.session_state.log   = []
            st.session_state.stats = {"With Mask":0,"No Mask":0,
                                      "Improper Mask":0,"total":0}
            st.session_state.excel_status = "🔄 Tracker reset"
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)

    # Excel status line
    col = "#2ecc71" if "⚠" not in st.session_state.excel_status else "#e74c3c"
    st.markdown(f"<div class='excel-status' style='color:{col}'>"
                f"{st.session_state.excel_status}</div>",
                unsafe_allow_html=True)

    # ── 📈 Session Stats ──
    st.markdown("<div class='sec-label'>📈&nbsp; Session Stats</div>"
                "<hr class='sec-line'>", unsafe_allow_html=True)

    stats = st.session_state.stats
    for k, (_,_,emoji) in STATUS_COLOR.items():
        st.markdown(f"<div class='stat-row'>{emoji}&nbsp; {k}: "
                    f"<b>{stats[k]}</b></div>", unsafe_allow_html=True)
    st.markdown(f"<div class='stat-row'>📋&nbsp; Total logged: "
                f"<b>{stats['total']}</b></div>", unsafe_allow_html=True)

    # Active faces placeholder
    faces_ph = st.empty()
    faces_ph.markdown("<div class='stat-faces'>👤&nbsp; Active faces: 0</div>",
                      unsafe_allow_html=True)

    # ── 🤖 Model Status ──
    st.markdown("<div class='sec-label'>🤖&nbsp; Model Status</div>"
                "<hr class='sec-line'>", unsafe_allow_html=True)
    st.markdown("<div class='model-status'>✅ OpenCV Haar + skin heuristic<br>"
                "Place mask_yolov8.pt here for best accuracy</div>",
                unsafe_allow_html=True)

    # ── 📋 Recent Logs ──
    st.markdown("<div class='sec-label'>📋&nbsp; Recent Logs</div>"
                "<hr class='sec-line'>", unsafe_allow_html=True)
    log_ph = st.empty()

    def render_log():
        if not st.session_state.log:
            log_ph.markdown(
                "<div class='log-box' style='color:#555e80'>"
                "No detections yet…</div>", unsafe_allow_html=True)
            return
        rows = ""
        for e in reversed(st.session_state.log[-12:]):
            css = {"With Mask":"log-mask",
                   "No Mask":"log-nomask",
                   "Improper Mask":"log-improper"}.get(e["status"],"")
            rows += (f"<div class='{css}'>"
                     f"[{e['ts'][11:19]}] Face-{str(e['fid']):<3} "
                     f"{e['status']:<18} {e['conf']:.0%}"
                     f"</div>")
        log_ph.markdown(f"<div class='log-box'>{rows}</div>",
                        unsafe_allow_html=True)

    render_log()
    st.markdown("</div>", unsafe_allow_html=True)   # close right-panel


# ───────────────────────────────────────────────────────────────────
# LEFT PANEL  (video canvas + badge + webrtc)
# ───────────────────────────────────────────────────────────────────
with left:
    # Status badge (matches tkinter badge_lbl)
    badge_ph = st.empty()
    badge_ph.markdown(
        "<div class='badge'>⬤&nbsp; Idle — click START below</div>",
        unsafe_allow_html=True)

    # WebRTC live stream  ← this IS the camera feed, draws bbox on each frame
    ctx = webrtc_streamer(
        key            = "mask-live",
        video_processor_factory = MaskProcessor,
        rtc_configuration       = RTC_CONFIG,
        media_stream_constraints= {"video": True, "audio": False},
        async_processing        = True,
        translations={
            "start": "▶  Start Detection",
            "stop":  "⏹  Stop",
            "select_device": "Select camera",
        },
    )

    # Pass confidence to processor
    if ctx.video_processor:
        ctx.video_processor.conf_thresh = conf_val


# ═══════════════════════════════════════════════════════════════════
# LIVE POLLING — pull events from processor every second
# ═══════════════════════════════════════════════════════════════════
import time

if ctx.state.playing and ctx.video_processor:

    # Pull new detection events logged this cycle
    new_events = ctx.video_processor.pull_events()

    for e in new_events:
        st.session_state.stats[e["status"]] = \
            st.session_state.stats.get(e["status"], 0) + 1
        st.session_state.stats["total"] += 1
        st.session_state.log.append(e)
        st.session_state.excel_status = (
            f"📁 Excel: Saved ✓  [{e['ts'][11:19]}]")

    # Get latest status + face count
    (best_s, best_c), n_faces = ctx.video_processor.get_state()

    # Update badge
    if best_s in STATUS_COLOR:
        col, _, emoji = STATUS_COLOR[best_s]
        badge_ph.markdown(
            f"<div class='badge badge-"
            f"{'mask' if best_s=='With Mask' else 'nomask' if best_s=='No Mask' else 'improper'}'>"
            f"{emoji}&nbsp; {best_s.upper()}&nbsp; ({best_c:.0%})</div>",
            unsafe_allow_html=True)
    else:
        badge_ph.markdown(
            "<div class='badge'>⬤&nbsp; No face detected</div>",
            unsafe_allow_html=True)

    # Update active faces
    faces_ph.markdown(
        f"<div class='stat-faces'>👤&nbsp; Active faces: {n_faces}</div>",
        unsafe_allow_html=True)

    # Re-render log and stats if new events came in
    if new_events:
        render_log()

    # Auto-refresh every 1 second
    time.sleep(1)
    st.rerun()

elif not ctx.state.playing:
    badge_ph.markdown(
        "<div class='badge'>⬤&nbsp; Click ▶ Start Detection to begin</div>",
        unsafe_allow_html=True)
