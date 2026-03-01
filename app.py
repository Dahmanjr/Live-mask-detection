"""
Live Mask Detection — WebRTC Real-time
=======================================
Uses streamlit-webrtc for true live camera streaming in browser.
Works on Streamlit Cloud — no photo upload needed.
"""
import streamlit as st
import cv2
import numpy as np
import io
import threading
from datetime import datetime
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from streamlit_webrtc import webrtc_streamer, VideoProcessorBase, RTCConfiguration
import av

# ── Page config ───────────────────────────────────────────────────
st.set_page_config(
    page_title="Live Mask Detection",
    page_icon="🎭",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;700;800&family=Space+Mono&display=swap');
html, body, [class*="css"] { font-family: 'Syne', sans-serif; }
.stApp { background: #0a0e1a; color: #e8eaf0; }
[data-testid="stSidebar"] {
    background: #10152a !important;
    border-right: 1px solid #1f2744;
}
.hero {
    font-size: 2.2rem; font-weight: 800;
    background: linear-gradient(135deg, #e94560, #f5a623);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}
.sub {
    font-family: 'Space Mono', monospace; font-size: .72rem;
    color: #555e80; letter-spacing: .12em; text-transform: uppercase;
}
.status-box {
    font-size: 1.4rem; font-weight: 700; padding: 16px 24px;
    border-radius: 12px; text-align: center; margin: 10px 0;
    font-family: 'Space Mono', monospace;
}
.s-mask     { background:#0d3320; color:#27ae60; border:1px solid #27ae60; }
.s-nomask   { background:#330d0d; color:#e74c3c; border:1px solid #e74c3c; }
.s-improper { background:#332a00; color:#f39c12; border:1px solid #f39c12; }
.s-none     { background:#13192e; color:#555e80; border:1px solid #1f2744; }
.log-row {
    font-family: 'Space Mono', monospace; font-size: .75rem;
    padding: 7px 12px; border-radius: 7px; margin-bottom: 4px;
    display: flex; justify-content: space-between;
}
.l-mask     { background:#0d3320; color:#27ae60; }
.l-nomask   { background:#330d0d; color:#e74c3c; }
.l-improper { background:#332a00; color:#f39c12; }
.live-pill {
    display:inline-block; background:#e74c3c; color:white;
    font-family:'Space Mono',monospace; font-size:.65rem;
    font-weight:700; letter-spacing:.15em; padding:4px 12px;
    border-radius:20px; animation:blink 1.2s infinite;
}
@keyframes blink { 0%,100%{opacity:1} 50%{opacity:.3} }
hr { border-color: #1f2744 !important; }
.stButton > button {
    background: #e94560 !important; color: white !important;
    border: none !important; border-radius: 8px !important;
    font-family: 'Syne', sans-serif !important; font-weight: 700 !important;
    width: 100%; padding: .5rem 1rem !important;
}
[data-testid="stMetricValue"] {
    font-family: 'Space Mono', monospace !important;
    font-size: 1.8rem !important; color: #e8eaf0 !important;
}
</style>
""", unsafe_allow_html=True)

# ── WebRTC ICE servers (needed for Streamlit Cloud NAT traversal) ──
RTC_CONFIG = RTCConfiguration({
    "iceServers": [
        {"urls": ["stun:stun.l.google.com:19302"]},
        {"urls": ["stun:stun1.l.google.com:19302"]},
    ]
})

# ═══════════════════════════════════════════════════════════════════
# FACE TRACKER
# ═══════════════════════════════════════════════════════════════════
class FaceTracker:
    MAX_GONE = 30
    MAX_DIST = 120

    def __init__(self):
        self.nid    = 0
        self.tracks = {}

    def _cen(self, x1, y1, x2, y2):
        return np.array([(x1+x2)//2, (y1+y2)//2], dtype=float)

    def update(self, dets):
        for t in self.tracks:
            self.tracks[t]["gone"] += 1
        for t in [k for k,v in self.tracks.items() if v["gone"] > self.MAX_GONE]:
            del self.tracks[t]

        if not dets:
            return []

        cents = [self._cen(*d[:4]) for d in dets]
        out   = []

        if not self.tracks:
            for i, d in enumerate(dets):
                tid = self._reg(cents[i], d)
                out.append((*d, tid, True))
            return out

        tids   = list(self.tracks.keys())
        tcents = np.array([self.tracks[t]["cen"] for t in tids])
        dists  = np.linalg.norm(
            tcents[:,None,:] - np.array(cents)[None,:,:], axis=2)

        used_t, used_d = set(), set()
        for _ in range(min(len(tids), len(dets))):
            if dists.size == 0: break
            ti, di = np.unravel_index(dists.argmin(), dists.shape)
            if dists[ti, di] > self.MAX_DIST: break
            tid = tids[ti]; used_t.add(ti); used_d.add(di)
            d   = dets[di]
            self.tracks[tid].update({"cen": cents[di], "gone": 0})
            prev   = self.tracks[tid]["last"]
            is_new = prev != d[4]
            if is_new:
                self.tracks[tid]["last"] = d[4]
            out.append((*d, tid, is_new))
            dists[ti,:] = 1e9; dists[:,di] = 1e9

        for di, d in enumerate(dets):
            if di not in used_d:
                tid = self._reg(cents[di], d)
                out.append((*d, tid, True))
        return out

    def _reg(self, cen, d):
        tid = self.nid; self.nid += 1
        self.tracks[tid] = {"cen": cen, "gone": 0, "last": d[4]}
        return tid


# ═══════════════════════════════════════════════════════════════════
# SKIN HEURISTIC
# ═══════════════════════════════════════════════════════════════════
_LO = np.array([0,  20,  70], np.uint8)
_HI = np.array([20, 255, 255], np.uint8)

def _skin(roi):
    if roi is None or roi.size == 0: return 0.0
    m = cv2.inRange(cv2.cvtColor(roi, cv2.COLOR_BGR2HSV), _LO, _HI)
    return cv2.countNonZero(m) / max(1, roi.shape[0] * roi.shape[1])

def classify(frame, x1, y1, x2, y2):
    fh = y2 - y1; ym = y1 + fh // 2
    if _skin(frame[ym:y2, x1:x2]) > 0.55:
        return "No Mask", 0.82
    if _skin(frame[y1+int(fh*.45) : y1+int(fh*.65), x1:x2]) > 0.40:
        return "Improper Mask", 0.74
    return "With Mask", 0.86


# ═══════════════════════════════════════════════════════════════════
# DRAWING
# ═══════════════════════════════════════════════════════════════════
BGRS  = {"With Mask":(0,210,0), "No Mask":(0,0,220), "Improper Mask":(0,165,255)}
EMOJI = {"With Mask":"✅", "No Mask":"❌", "Improper Mask":"⚠️"}
CSS   = {"With Mask":"s-mask", "No Mask":"s-nomask", "Improper Mask":"s-improper"}
LCSS  = {"With Mask":"l-mask", "No Mask":"l-nomask", "Improper Mask":"l-improper"}

def draw_boxes(frame, tracked):
    for (x1,y1,x2,y2,status,conf,fid,is_new) in tracked:
        bgr = BGRS.get(status, (150,150,150))
        cv2.rectangle(frame, (x1,y1), (x2,y2), bgr, 3)
        lbl = f" Face-{fid} | {status}  {conf:.0%}"
        (tw,th),_ = cv2.getTextSize(lbl, cv2.FONT_HERSHEY_SIMPLEX, 0.65, 2)
        cv2.rectangle(frame, (x1,y1-th-16), (x1+tw+4,y1), bgr, -1)
        cv2.putText(frame, lbl, (x1+2, y1-6),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255,255,255), 2, cv2.LINE_AA)
        if is_new:
            cv2.putText(frame, "● LOGGED", (x1, y2+22),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.55, bgr, 2, cv2.LINE_AA)
    return frame


# ═══════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════
def build_excel(log):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Detection Log"
    headers = ["#","Timestamp","Date","Time",
               "Face ID","Status","Confidence (%)","Notes"]
    hf   = PatternFill("solid", fgColor="1F4E79")
    hfnt = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hf; cell.font = hfnt
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for i, w in enumerate([5,22,12,10,9,18,16,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    fills = {"With Mask":    ("C6EFCE","276221"),
             "No Mask":      ("FFC7CE","9C0006"),
             "Improper Mask":("FFEB9C","9C5700")}
    for idx, e in enumerate(log, 1):
        bg,fg = fills.get(e["status"], ("FFFFFF","000000"))
        fill  = PatternFill("solid", fgColor=bg)
        font  = Font(color=fg, name="Arial")
        vals  = [idx, e["ts"], e["ts"][:10], e["ts"][11:19],
                 f"Face-{e['fid']}", e["status"],
                 round(e["conf"]*100,1), "Live detection"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=idx+1, column=c, value=v)
            cell.fill=fill; cell.font=font
            cell.alignment=Alignment(horizontal="center")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════
# VIDEO PROCESSOR — runs on every live frame
# ═══════════════════════════════════════════════════════════════════
class MaskProcessor(VideoProcessorBase):
    def __init__(self):
        self.detector = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        self.tracker  = FaceTracker()
        self.conf_thresh = 0.45
        self.lock     = threading.Lock()

        # Shared result — read by main thread for logging
        self.last_tracked  = []
        self.new_events    = []   # list of (fid, status, conf) to log

    def recv(self, frame: av.VideoFrame) -> av.VideoFrame:
        img = frame.to_ndarray(format="bgr24")

        # Detect faces
        gray  = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = self.detector.detectMultiScale(
                    gray, scaleFactor=1.1, minNeighbors=5, minSize=(50,50))

        dets = []
        for (fx, fy, fw, fh) in faces:
            status, conf = classify(img, fx, fy, fx+fw, fy+fh)
            if conf >= self.conf_thresh:
                dets.append((fx, fy, fx+fw, fy+fh, status, conf))

        tracked = self.tracker.update(dets)

        # Collect new log events
        with self.lock:
            self.last_tracked = tracked
            for item in tracked:
                if item[7]:   # is_new_status
                    self.new_events.append({
                        "fid":    item[6],
                        "status": item[4],
                        "conf":   item[5],
                        "ts":     datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    })

        # Draw on frame
        annotated = draw_boxes(img.copy(), tracked)
        return av.VideoFrame.from_ndarray(annotated, format="bgr24")

    def get_and_clear_events(self):
        with self.lock:
            evts = self.new_events.copy()
            self.new_events.clear()
            return evts

    def get_tracked(self):
        with self.lock:
            return self.last_tracked.copy()


# ═══════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════
for k, v in [("log",[]),
              ("stats",{"With Mask":0,"No Mask":0,"Improper Mask":0})]:
    if k not in st.session_state:
        st.session_state[k] = v


# ═══════════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════════
st.markdown("""
<div style='padding:1rem 0 .5rem'>
  <div class='hero'>🎭 Live Mask Detection</div>
  <div class='sub'>Real-time WebRTC · Per-person tracking · Excel export</div>
</div><hr>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### ⚙️ Settings")
    conf_val = st.slider("Confidence threshold", 0.10, 0.95, 0.45, 0.05)

    st.markdown("---")
    st.markdown("### 📊 Session Stats")
    stat_ph = st.empty()

    def render_stats():
        c1,c2,c3 = stat_ph.columns(3)
        c1.metric("✅ Mask",     st.session_state.stats["With Mask"])
        c2.metric("❌ No Mask",  st.session_state.stats["No Mask"])
        c3.metric("⚠️ Improper", st.session_state.stats["Improper Mask"])

    render_stats()

    st.markdown("---")
    st.markdown("### ℹ️ How it works")
    st.info(
        "**Smart tracking:**\n\n"
        "- Each face gets a **Face ID**\n"
        "- Logged **once** on first detection\n"
        "- Logged again **only if** status changes\n"
        "  *(mask removed or put on)*"
    )

    if st.button("🔄 Reset All"):
        st.session_state.log   = []
        st.session_state.stats = {"With Mask":0,"No Mask":0,"Improper Mask":0}
        st.rerun()

    st.markdown("---")
    if st.session_state.log:
        st.download_button(
            "📥 Download Excel Log",
            data=build_excel(st.session_state.log),
            file_name=f"mask_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    else:
        st.caption("Excel download appears after first detection.")


# ═══════════════════════════════════════════════════════════════════
# MAIN LAYOUT
# ═══════════════════════════════════════════════════════════════════
left, right = st.columns([3, 2], gap="large")

with left:
    st.markdown("### 📷 Live Camera")
    st.markdown(
        "<div class='live-pill'>● LIVE</div>",
        unsafe_allow_html=True)

    # ── WebRTC streamer ──────────────────────────────────────────
    ctx = webrtc_streamer(
        key="mask-detection",
        video_processor_factory=MaskProcessor,
        rtc_configuration=RTC_CONFIG,
        media_stream_constraints={"video": True, "audio": False},
        async_processing=True,
    )

    # Pass confidence threshold to processor
    if ctx.video_processor:
        ctx.video_processor.conf_thresh = conf_val

    status_ph = st.empty()

with right:
    st.markdown("### 📋 Detection Log")
    log_ph = st.empty()

    def render_log():
        if not st.session_state.log:
            log_ph.info("🎥 Start camera to begin detection!"); return
        html = ""
        for e in reversed(st.session_state.log[-25:]):
            lc = LCSS.get(e["status"], "")
            em = EMOJI.get(e["status"], "")
            html += (f"<div class='log-row {lc}'>"
                     f"<span>Face-{e['fid']}&nbsp;&nbsp;{em} {e['status']}</span>"
                     f"<span>{e['ts'][11:19]}&nbsp;{e['conf']:.0%}</span>"
                     f"</div>")
        log_ph.markdown(html, unsafe_allow_html=True)

    render_log()

    st.markdown("### 👤 Tracking")
    track_ph = st.empty()
    track_ph.info("👤 0 face(s) in frame")


# ═══════════════════════════════════════════════════════════════════
# LIVE UPDATE LOOP — polls processor for new events
# ═══════════════════════════════════════════════════════════════════
if ctx.state.playing and ctx.video_processor:
    import time

    # Drain new detection events from the video processor
    new_events = ctx.video_processor.get_and_clear_events()

    for e in new_events:
        st.session_state.stats[e["status"]] = \
            st.session_state.stats.get(e["status"], 0) + 1
        st.session_state.log.append(e)

    # Update status badge from latest tracked faces
    tracked = ctx.video_processor.get_tracked()
    n_faces = len(ctx.video_processor.tracker.tracks)
    track_ph.info(f"👤 {n_faces} face(s) in frame")

    if tracked:
        best   = max(tracked, key=lambda x: x[5])
        bstat  = best[4]
        bconf  = best[5]
        status_ph.markdown(
            f"<div class='status-box {CSS.get(bstat,'s-none')}'>"
            f"{EMOJI.get(bstat,'')}  {bstat.upper()}  ({bconf:.0%})</div>",
            unsafe_allow_html=True)
    else:
        status_ph.markdown(
            "<div class='status-box s-none'>⬤  No face detected</div>",
            unsafe_allow_html=True)

    render_stats()
    render_log()

    # Auto-refresh every second to pull new detections
    time.sleep(1)
    st.rerun()

elif not ctx.state.playing:
    status_ph.markdown(
        "<div class='status-box s-none'>"
        "⬤  Click START above to begin live detection</div>",
        unsafe_allow_html=True)
